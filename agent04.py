"""
Vertex AI Search データストアを参照するADKエージェント
Discovery Engine APIを直接使用してデータストアを検索し、GCSからファイル内容を取得します
"""

from google.adk.agents import LlmAgent
from google.adk.tools import FunctionTool
from google.cloud import discoveryengine_v1 as discoveryengine
from google.cloud import storage
from typing import Any
import logging
import io

# ログ設定
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# プロジェクト設定
PROJECT_ID = "agent-vi-473112"
DATASTORE_ID = "adk-test_1769691409159"
DATASTORE_REGION = "global"


async def search_datastore(query: str) -> dict[str, Any]:
    """
    Vertex AI Searchのデータストアを検索します。
    
    Args:
        query: 検索クエリ文字列
        
    Returns:
        検索結果を含む辞書
    """
    logger.info(f"=== search_datastore called with query: {query} ===")
    
    try:
        client = discoveryengine.SearchServiceClient()
        serving_config = f"projects/{PROJECT_ID}/locations/{DATASTORE_REGION}/dataStores/{DATASTORE_ID}/servingConfigs/default_search"
        
        request = discoveryengine.SearchRequest(
            serving_config=serving_config,
            query=query,
            page_size=10,
            content_search_spec=discoveryengine.SearchRequest.ContentSearchSpec(
                snippet_spec=discoveryengine.SearchRequest.ContentSearchSpec.SnippetSpec(
                    return_snippet=True,
                ),
            ),
        )
        
        response = client.search(request)
        
        results = []
        for result in response.results:
            doc = result.document
            doc_data = {"id": doc.id}
            
            if doc.struct_data:
                for key, value in doc.struct_data.items():
                    doc_data[key] = str(value)[:500]
                    
            if doc.derived_struct_data:
                derived = dict(doc.derived_struct_data)
                if "snippets" in derived:
                    snippets_list = []
                    for snippet in derived["snippets"]:
                        if hasattr(snippet, 'snippet'):
                            snippets_list.append(snippet.snippet)
                        elif isinstance(snippet, dict) and 'snippet' in snippet:
                            snippets_list.append(snippet['snippet'])
                        else:
                            snippets_list.append(str(snippet))
                    doc_data["snippets"] = snippets_list
                if "link" in derived:
                    doc_data["link"] = str(derived["link"])
                if "title" in derived:
                    doc_data["title"] = str(derived["title"])
                    
            results.append(doc_data)
        
        logger.info(f"Total results found: {len(results)}")
        
        return {
            "success": True,
            "query": query,
            "total_results": len(results),
            "results": results
        }
        
    except Exception as e:
        logger.error(f"Search error: {e}")
        return {"success": False, "error": str(e), "query": query}


async def get_document_content(document_id: str) -> dict[str, Any]:
    """
    ドキュメントIDからドキュメントの詳細とGCS上のファイル内容を取得します。
    
    Args:
        document_id: ドキュメントID
        
    Returns:
        ドキュメントの内容を含む辞書
    """
    logger.info(f"=== get_document_content called with id: {document_id} ===")
    
    try:
        # ドキュメント情報を取得
        doc_client = discoveryengine.DocumentServiceClient()
        doc_name = f"projects/{PROJECT_ID}/locations/{DATASTORE_REGION}/dataStores/{DATASTORE_ID}/branches/default_branch/documents/{document_id}"
        
        doc = doc_client.get_document(name=doc_name)
        logger.info(f"Document retrieved: {doc.name}")
        
        result = {
            "id": doc.id,
            "name": doc.name,
        }
        
        # GCSリンクを取得
        gcs_uri = None
        if doc.content and doc.content.uri:
            gcs_uri = doc.content.uri
            result["gcs_uri"] = gcs_uri
            logger.info(f"GCS URI: {gcs_uri}")
        
        # GCSからファイルを読み込む
        if gcs_uri and gcs_uri.startswith("gs://"):
            content = await _read_gcs_file(gcs_uri)
            if content:
                result["file_content"] = content
        
        return {"success": True, "document": result}
        
    except Exception as e:
        logger.error(f"Get document error: {e}")
        return {"success": False, "error": str(e)}


async def _read_gcs_file(gcs_uri: str) -> dict[str, Any] | None:
    """GCSからファイルを読み込む"""
    try:
        # gs://bucket/path 形式をパース
        parts = gcs_uri.replace("gs://", "").split("/", 1)
        bucket_name = parts[0]
        blob_name = parts[1] if len(parts) > 1 else ""
        
        logger.info(f"Reading from GCS: bucket={bucket_name}, blob={blob_name}")
        
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        
        # ファイルをダウンロード
        content = blob.download_as_bytes()
        
        # Excelファイルの場合
        if blob_name.endswith('.xlsx') or blob_name.endswith('.xls'):
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            
            sheets_data = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = []
                for row in sheet.iter_rows(max_row=50):  # 最大50行
                    row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                    if any(row_data):  # 空行はスキップ
                        rows.append(row_data)
                sheets_data[sheet_name] = rows
            
            return {
                "type": "excel",
                "filename": blob_name,
                "sheets": sheets_data
            }
        
        # テキストファイルの場合
        elif blob_name.endswith('.txt') or blob_name.endswith('.csv'):
            text = content.decode('utf-8')
            return {
                "type": "text",
                "filename": blob_name,
                "content": text[:5000]  # 最大5000文字
            }
        
        # その他のファイル
        else:
            return {
                "type": "binary",
                "filename": blob_name,
                "size_bytes": len(content)
            }
            
    except Exception as e:
        logger.error(f"GCS read error: {e}")
        return None


async def list_all_documents() -> dict[str, Any]:
    """
    データストア内の全ドキュメントを一覧表示します。
    
    Returns:
        ドキュメント一覧
    """
    logger.info("=== list_all_documents called ===")
    
    try:
        client = discoveryengine.DocumentServiceClient()
        parent = f"projects/{PROJECT_ID}/locations/{DATASTORE_REGION}/dataStores/{DATASTORE_ID}/branches/default_branch"
        
        docs = client.list_documents(parent=parent)
        
        results = []
        for doc in docs:
            doc_info = {
                "id": doc.id,
                "name": doc.name,
            }
            if doc.content and doc.content.uri:
                doc_info["gcs_uri"] = doc.content.uri
            results.append(doc_info)
        
        logger.info(f"Total documents: {len(results)}")
        
        return {
            "success": True,
            "total": len(results),
            "documents": results
        }
        
    except Exception as e:
        logger.error(f"List documents error: {e}")
        return {"success": False, "error": str(e)}


# ツールを登録
search_tool = FunctionTool(func=search_datastore)
get_content_tool = FunctionTool(func=get_document_content)
list_docs_tool = FunctionTool(func=list_all_documents)

# エージェント定義
root_agent = LlmAgent(
    name="vertex_search_agent",
    model="gemini-2.0-flash",
    tools=[search_tool, get_content_tool, list_docs_tool],
    instruction="""あなたはVertex AI Searchを使用してドキュメントを検索し、質問に回答するアシスタントです。

## 使用するツール
1. **search_datastore**: キーワードでデータストアを検索します
2. **get_document_content**: ドキュメントIDを指定してファイルの中身を取得します
3. **list_all_documents**: データストア内の全ドキュメントを一覧表示します

## 基本的なワークフロー
1. まず「何が入っているか」を聞かれたら `list_all_documents` を使って一覧を表示
2. キーワードで検索する場合は `search_datastore` を使用（シンプルな1〜3語のキーワードで）
3. ファイルの中身を見たい場合は `get_document_content` でドキュメントIDを指定して取得

## 注意点
- 検索クエリはシンプルなキーワードで（例：「会計」「売上」「データ」）
- ドキュメントIDは検索結果やリストから取得できます
- Excelファイルの場合、シート名とセルの内容が返されます

日本語で回答してください。
""",
    description="Vertex AI Searchデータストアを検索し、ファイル内容を取得するエージェント",
)
