"""
Excel出力用ツール

BigQueryから取得したデータをExcelファイルとしてArtifactsに保存する
"""
import io
from typing import Any
import google.genai.types as types
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _extract_value(val: Any) -> Any:
    """
    BigQueryの結果形式から実際の値を抽出する
    
    BigQueryは {'v': value} 形式で値を返すことがある
    """
    if isinstance(val, dict):
        if 'v' in val:
            return _extract_value(val['v'])
        return str(val)
    elif isinstance(val, list):
        return [_extract_value(v) for v in val]
    else:
        return val


def _normalize_bq_data(data: Any) -> list[dict[str, Any]]:
    """
    BigQueryの様々な結果形式を統一された辞書リストに変換する
    """
    if not data:
        return []
    
    # 既に正規化された形式の場合
    if isinstance(data, list) and len(data) > 0:
        first_row = data[0]
        
        # 標準的な辞書リスト形式
        if isinstance(first_row, dict) and 'f' not in first_row and 'v' not in first_row:
            normalized = []
            for row in data:
                normalized_row = {}
                for key, val in row.items():
                    normalized_row[key] = _extract_value(val)
                normalized.append(normalized_row)
            return normalized
        
        # BigQuery行形式: [{"f": [{"v": val1}, {"v": val2}]}, ...]
        if isinstance(first_row, dict) and 'f' in first_row:
            headers = [f"column_{i}" for i in range(len(first_row['f']))]
            normalized = []
            for row in data:
                fields = row.get('f', [])
                normalized_row = {}
                for i, field in enumerate(fields):
                    col_name = headers[i] if i < len(headers) else f"column_{i}"
                    normalized_row[col_name] = _extract_value(field)
                normalized.append(normalized_row)
            return normalized
    
    # 辞書形式（schema + rows）
    if isinstance(data, dict):
        rows = data.get('rows', [])
        schema = data.get('schema', {})
        fields = schema.get('fields', [])
        
        headers = [f.get('name', f'column_{i}') for i, f in enumerate(fields)]
        
        if not headers and rows:
            first_row = rows[0] if rows else {}
            if isinstance(first_row, dict) and 'f' in first_row:
                headers = [f"column_{i}" for i in range(len(first_row['f']))]
            elif isinstance(first_row, list):
                headers = [f"column_{i}" for i in range(len(first_row))]
        
        normalized = []
        for row in rows:
            if isinstance(row, dict) and 'f' in row:
                fields_data = row['f']
                normalized_row = {}
                for i, field in enumerate(fields_data):
                    col_name = headers[i] if i < len(headers) else f"column_{i}"
                    normalized_row[col_name] = _extract_value(field)
                normalized.append(normalized_row)
            elif isinstance(row, list):
                normalized_row = {}
                for i, val in enumerate(row):
                    col_name = headers[i] if i < len(headers) else f"column_{i}"
                    normalized_row[col_name] = _extract_value(val)
                normalized.append(normalized_row)
            elif isinstance(row, dict):
                normalized_row = {k: _extract_value(v) for k, v in row.items()}
                normalized.append(normalized_row)
        
        return normalized
    
    return []


async def export_to_excel(
    data: list[dict[str, Any]],
    filename: str,
    sheet_name: str = "Sheet1",
    tool_context: Any = None
) -> dict[str, Any]:
    """
    データをExcelファイルとして保存し、Artifactとして出力する
    """
    normalized_data = _normalize_bq_data(data)
    
    if not normalized_data:
        return {
            "success": False,
            "error": "データが空です。Excelファイルを作成できません。"
        }
    
    if not filename.endswith('.xlsx'):
        filename = f"{filename}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # スタイル定義
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ヘッダー行
    headers = list(normalized_data[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # データ行
    for row_idx, row_data in enumerate(normalized_data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            if isinstance(value, (list, dict)):
                value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # 列幅調整
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    # フィルター設定
    ws.auto_filter.ref = ws.dimensions
    
    # バイトストリームに保存
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()
    excel_buffer.close()
    
    # Artifactとして保存
    if tool_context:
        try:
            excel_artifact = types.Part.from_bytes(
                data=excel_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            version = await tool_context.save_artifact(
                filename=filename,
                artifact=excel_artifact
            )
            return {
                "success": True,
                "filename": filename,
                "rows": len(normalized_data),
                "columns": len(headers),
                "version": version,
                "size_bytes": len(excel_bytes),
                "message": f"Excelファイル '{filename}' を保存しました（{len(normalized_data)}行 x {len(headers)}列）"
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"Artifact保存エラー: {str(e)}",
                "filename": filename
            }
    else:
        return {
            "success": False,
            "error": "ToolContextが提供されていません。",
            "filename": filename
        }


async def list_saved_files(tool_context: Any = None) -> dict[str, Any]:
    """保存済みのArtifactファイル一覧を取得する"""
    if not tool_context:
        return {"success": False, "error": "ToolContextが提供されていません。"}
    
    try:
        files = await tool_context.list_artifacts()
        return {
            "success": True,
            "files": files if files else [],
            "count": len(files) if files else 0
        }
    except Exception as e:
        return {"success": False, "error": f"ファイル一覧取得エラー: {str(e)}"}
